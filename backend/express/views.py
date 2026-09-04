from decimal import Decimal, InvalidOperation

from django.db.models import Count, Q, Sum
from django.utils import timezone
from rest_framework import serializers, viewsets
from rest_framework.decorators import action, api_view, permission_classes, throttle_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from loko.throttling import PublicReadThrottle, PublicWriteThrottle

from drf_spectacular.types import OpenApiTypes
from drf_spectacular.utils import (
    OpenApiParameter,
    extend_schema,
    extend_schema_view,
    inline_serializer,
)

from accounts.permissions import (
    DenyOperatorOrDirector,
    IsAdmin,
    SalesAccess,
    StockAccess,
    WarehouseAccess,
    WarehouseItemAccess,
    WorkflowAccess,
)
from finance.models import Account, AppSettings, Branch
from .models import Client, ClientPrice, EmployeeRating, Sale, WarehouseItem, WarehouseOrder, WarehouseStock
from .workflow import build_stock, build_workflow
from .serializers import (
    ClientPriceSerializer,
    ClientSerializer,
    OperatorSaleSerializer,
    PublicIntakeSerializer,
    PublicRateSerializer,
    SaleSerializer,
    WarehouseItemSerializer,
    WarehouseNotFoundSerializer,
    WarehouseOrderSerializer,
    WarehouseReceiveSerializer,
    WarehouseStatusSerializer,
    WarehouseStockSerializer,
)

ZERO = Decimal("0.00")


@extend_schema_view(
    list=extend_schema(
        parameters=[
            OpenApiParameter("from", OpenApiTypes.DATE, description="Начало периода"),
            OpenApiParameter("to", OpenApiTypes.DATE, description="Конец периода"),
            OpenApiParameter("payment", OpenApiTypes.STR, enum=["all", "cash", "noncash"], description="Вид оплаты"),
            OpenApiParameter("search", OpenApiTypes.STR, description="Поиск по коду клиента"),
            OpenApiParameter("branch", OpenApiTypes.INT, description="Филиал Express (id)"),
        ]
    )
)
class SaleViewSet(viewsets.ModelViewSet):
    serializer_class = SaleSerializer
    # Managers/admins: full access. Operators («Сотрудник»): create + quote +
    # the minimal account picker only (enforced per-action by SalesAccess).
    permission_classes = [SalesAccess]

    def get_permissions(self):
        # Экспорт всей таблицы продаж (с финансовыми полями) — только администратор.
        if self.action == "export":
            return [IsAdmin()]
        return [SalesAccess()]

    def get_serializer_class(self):
        # Операторам — узкий сериализатор без финансовых полей: ни себестоимость/
        # маржа/ставки в ответе на create, ни возможность задать их через тело.
        if getattr(self.request.user, "is_operator", False):
            return OperatorSaleSerializer
        return SaleSerializer

    def get_queryset(self):
        qs = Sale.objects.select_related("account").all()
        params = self.request.query_params
        date_from = params.get("from")
        date_to = params.get("to")
        payment = params.get("payment")
        search = params.get("search")
        branch = params.get("branch")
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if payment == "cash":
            qs = qs.filter(account__kind="CASH")
        elif payment == "noncash":
            qs = qs.filter(account__kind="BANK")
        if search:
            qs = qs.filter(client_code__icontains=search)
        if branch:
            qs = qs.filter(branch=branch)
        return qs

    def perform_create(self, serializer):
        # Прямая продажа Express — только кассир/админ (Sales.jsx). Оператор продажи
        # НЕ создаёт: он формирует складскую заявку, а продажа рождается при
        # оприходовании складовщиком (WarehouseItem.receive → Sale). Филиал —
        # выбранный в форме, иначе филиал по умолчанию.
        branch = serializer.validated_data.get("branch") or Branch.resolve_default()
        serializer.save(created_by=self.request.user, branch=branch)

    @extend_schema(responses=OpenApiTypes.OBJECT)
    @action(detail=False, methods=["get"])
    def summary(self, request):
        """Aggregated totals over the current filter (for dashboard cards)."""
        qs = self.get_queryset()
        agg = qs.aggregate(
            count=Count("id"),
            revenue=Sum("price_som"),
            paid=Sum("paid_som"),
            cost=Sum("cost_som"),
            margin=Sum("margin_som"),
            weight=Sum("weight_kg"),
        )
        revenue = agg["revenue"] or ZERO
        paid = agg["paid"] or ZERO
        return Response(
            {
                "count": agg["count"] or 0,
                "revenue": revenue,                 # начисление
                "paid": paid,                       # оплата
                "receivable": revenue - paid,       # дебиторка
                "cost": agg["cost"] or ZERO,
                "margin": agg["margin"] or ZERO,
                "weight": agg["weight"] or ZERO,
            }
        )

    @extend_schema(
        parameters=[
            OpenApiParameter("from", OpenApiTypes.DATE, description="Начало периода"),
            OpenApiParameter("to", OpenApiTypes.DATE, description="Конец периода"),
            OpenApiParameter("payment", OpenApiTypes.STR, enum=["all", "cash", "noncash"], description="Вид оплаты"),
            OpenApiParameter("search", OpenApiTypes.STR, description="Поиск по коду клиента"),
            OpenApiParameter("branch", OpenApiTypes.INT, description="Филиал Express (id)"),
        ],
        responses=OpenApiTypes.BINARY,
    )
    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        """Выгрузка продаж Express (по текущим фильтрам) в Excel — только администратор.

        Полная таблица с финансами: себестоимость, маржа, дебиторка + итоговая строка.
        Фильтры (даты/оплата/поиск/филиал) — те же, что и в списке (``get_queryset``).
        """
        qs = self.get_queryset().select_related("account", "branch").order_by("-date", "-id")
        return self._export_xlsx(qs)

    @staticmethod
    def _export_xlsx(qs):
        """Полная выгрузка продаж Express в .xlsx (openpyxl) с итоговой строкой."""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Продажи Express"
        headers = [
            "Дата", "Код клиента", "Режим", "Вес, кг", "Кол-во", "Счёт", "Филиал",
            "Начислено, сом", "Оплачено, сом", "Дебиторка, сом",
            "Себестоимость, сом", "Маржа, сом", "Дата оплаты",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        t_price = t_paid = t_recv = t_cost = t_margin = ZERO
        for s in qs:
            price = s.price_som or ZERO
            paid = s.paid_som or ZERO
            recv = s.receivable_som
            cost = s.cost_som or ZERO
            margin = s.margin_som or ZERO
            t_price += price
            t_paid += paid
            t_recv += recv
            t_cost += cost
            t_margin += margin
            ws.append([
                s.date.strftime("%d.%m.%Y") if s.date else "",
                s.client_code,
                s.get_amount_mode_display(),
                float(s.weight_kg) if s.weight_kg is not None else None,
                s.places,
                s.account.name if s.account_id else "",
                s.branch.name if s.branch_id else "—",
                float(price), float(paid), float(recv), float(cost), float(margin),
                s.payment_date.strftime("%d.%m.%Y") if s.payment_date else "",
            ])

        ws.append([
            "ИТОГО", "", "", "", "", "", "",
            float(t_price), float(t_paid), float(t_recv), float(t_cost), float(t_margin), "",
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

        for i, width in enumerate((12, 16, 16, 10, 8, 16, 26, 15, 14, 14, 16, 14, 12)):
            ws.column_dimensions[ws.cell(row=1, column=i + 1).column_letter].width = width

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="sales-express.xlsx"'
        wb.save(resp)
        return resp

    @extend_schema(
        parameters=[
            OpenApiParameter("from", OpenApiTypes.DATE, description="Начало периода"),
            OpenApiParameter("to", OpenApiTypes.DATE, description="Конец периода"),
        ],
        responses=OpenApiTypes.OBJECT,
    )
    @action(detail=False, methods=["get"], url_path="weight-summary")
    def weight_summary(self, request):
        """Общий вес (фактический + расчётный) за отдельный период.

        Виджет с собственным фильтром дат — независим от основного фильтра списка
        (``from``/``to`` здесь свои). Суммирует явно введённый вес, а для продаж
        «прямой суммой» без веса — расчётный вес из суммы, ровно как значения «≈»
        в таблице (см. ``Sale.est_weight_kg``): цена ÷ (цена_за_кг_$ × курс_$).
        """
        qs = Sale.objects.all()
        date_from = request.query_params.get("from")
        date_to = request.query_params.get("to")
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        # est_weight_kg — свойство модели (не поле БД), поэтому агрегируем в Python
        # по тем же полям, что и в таблице → карточка сходится с колонкой «Вес».
        qs = qs.only("weight_kg", "price_som", "price_per_kg_usd", "usd_rate_som")
        actual = ZERO  # из явно введённого веса
        est = ZERO     # расчётный (из суммы) — продажи «прямой суммой» без веса
        count = 0
        for sale in qs.iterator():
            count += 1
            w = sale.est_weight_kg
            if w is None:
                continue
            if sale.weight_kg is not None:
                actual += w
            else:
                est += w
        return Response(
            {
                "count": count,
                "weight": actual + est,   # общий вес: факт + расчётный
                "weight_actual": actual,
                "weight_est": est,
            }
        )

    @extend_schema(
        request=inline_serializer(
            "SaleQuoteRequest",
            {
                "weight_kg": serializers.DecimalField(max_digits=10, decimal_places=3),
                "client_code": serializers.CharField(required=False),
            },
        ),
        responses=OpenApiTypes.OBJECT,
    )
    @action(detail=False, methods=["post"], url_path="quote")
    def quote(self, request):
        """Live price/cost/margin preview without persisting a sale.

        Если передан ``client_code`` со спец-ценой — итог считается по цене клиента
        (саму цену в ответе оператору не раскрываем — только итоговую стоимость)."""
        cfg = AppSettings.load()
        try:
            weight = Decimal(str(request.data.get("weight_kg", "0")))
        except (TypeError, ValueError, InvalidOperation):
            weight = ZERO
        # Отсекаем не-числа, отрицательные и абсурдно большие значения (иначе
        # quantize переполняется → 500). Верх — ёмкость поля weight_kg.
        if not weight.is_finite() or weight < 0 or weight > Decimal("9999999.999"):
            weight = ZERO
        # Спец-цена клиента (если есть) — иначе цена по умолчанию из Настроек.
        code = (request.data.get("client_code") or "").strip()
        unit = (
            ClientPrice.objects.filter(client_code=code).values_list("price_per_kg_som", flat=True).first()
            if code else None
        )
        if unit is not None:
            price = (weight * unit).quantize(ZERO)
        else:
            price = (weight * cfg.price_per_kg_usd * cfg.usd_rate_som).quantize(ZERO)
        cost = (weight * cfg.base_cost_per_kg_som).quantize(ZERO)
        # Операторам («Сотрудник») отдаём ТОЛЬКО общую стоимость — без себестоимости,
        # маржи и ставок (даже на уровне API, чтобы их не было видно в devtools).
        if getattr(request.user, "is_operator", False):
            return Response({"weight_kg": weight, "price_som": price})
        return Response(
            {
                "weight_kg": weight,
                "price_per_kg_usd": cfg.price_per_kg_usd,
                "usd_rate_som": cfg.usd_rate_som,
                "cost_per_kg_som": cfg.base_cost_per_kg_som,
                "price_som": price,
                "cost_som": cost,
                "margin_som": price - cost,
            }
        )

    @extend_schema(responses=OpenApiTypes.OBJECT)
    @action(detail=False, methods=["get"], url_path="accounts")
    def express_accounts(self, request):
        """Minimal Express-account picker for the sale form.

        Returns only id/name/kind (NO balances) so the «Сотрудник» role can
        choose where a sale is credited without touching the finance-laden
        /accounts/ endpoint. Sales may only land on Express accounts in сом.
        """
        qs = Account.objects.filter(
            module="EXPRESS", currency="KGS", is_active=True
        ).order_by("name")
        return Response([{"id": a.id, "name": a.name, "kind": a.kind} for a in qs])

    @extend_schema(responses=OpenApiTypes.OBJECT)
    @action(detail=False, methods=["get"], url_path="branches")
    def branches(self, request):
        """Минимальный пикер филиалов Express для формы продажи (id/name активных).

        Доступен и роли «Сотрудник»: его выбор всё равно жёстко фиксируется его
        филиалом на сервере (perform_create). Балансов/финансов не раскрывает.
        """
        from finance.models import Branch
        qs = Branch.objects.filter(is_active=True).order_by("name")
        return Response([{"id": b.id, "name": b.name} for b in qs])

    @extend_schema(
        parameters=[
            OpenApiParameter(
                "export", OpenApiTypes.STR, enum=["xlsx"],
                description="Формат выгрузки: xlsx — файл Excel; иначе JSON",
            )
        ],
        responses=OpenApiTypes.OBJECT,
    )
    @action(detail=False, methods=["get"], url_path="mine")
    def mine(self, request):
        """Свои продажи сотрудника за ТЕКУЩИЙ месяц (новые сверху).

        Возвращает продажи, созданные этим пользователем (``created_by``), с начала
        текущего календарного месяца (по локальному времени). ``?export=xlsx``
        отдаёт файл Excel. Узкий OperatorSaleSerializer не раскрывает
        себестоимость/маржу (как и при создании).
        """
        now_local = timezone.localtime(timezone.now())
        month_start = now_local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        qs = (
            Sale.objects.filter(created_by=request.user, created_at__gte=month_start)
            .select_related("account")
            .prefetch_related("warehouse_orders")
            .order_by("-created_at")
        )
        if request.query_params.get("export") == "xlsx":
            return self._mine_xlsx(qs)
        total = qs.aggregate(s=Sum("price_som"))["s"] or ZERO
        data = self.get_serializer(qs, many=True).data
        return Response({"count": qs.count(), "total_som": total, "results": data})

    @staticmethod
    def _mine_xlsx(qs):
        """Выгрузка «моих продаж за месяц» в .xlsx (openpyxl)."""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Мои продажи"
        ws.append(["Дата", "Время", "Код клиента", "Режим", "Вес, кг", "Кол-во", "Сумма, сом", "Счёт"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for s in qs:
            dt = s.created_at
            if timezone.is_aware(dt):
                dt = timezone.localtime(dt)
            ws.append([
                dt.strftime("%d.%m.%Y"),
                dt.strftime("%H:%M"),
                s.client_code,
                s.get_amount_mode_display(),
                float(s.weight_kg) if s.weight_kg is not None else None,
                s.places,
                float(s.price_som or 0),
                s.account.name,
            ])
        for col, width in zip("ABCDEFGH", (12, 8, 16, 16, 10, 8, 12, 18)):
            ws.column_dimensions[col].width = width

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="my-sales.xlsx"'
        wb.save(resp)
        return resp


@extend_schema_view(
    list=extend_schema(
        parameters=[OpenApiParameter("client_code", OpenApiTypes.STR, description="Точный код клиента (для подстановки цены)")]
    )
)
class ClientPriceViewSet(viewsets.ModelViewSet):
    """Индивидуальные цены за кг по клиентам (Express).

    Менеджер/админ управляют; операторы и директора — без доступа. Создание —
    «upsert»: если для кода клиента цена уже есть, обновляем её (не плодим дубли)."""

    serializer_class = ClientPriceSerializer
    permission_classes = [DenyOperatorOrDirector]

    def get_queryset(self):
        qs = ClientPrice.objects.all()
        code = self.request.query_params.get("client_code")
        if code:
            qs = qs.filter(client_code=code.strip())
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def create(self, request, *args, **kwargs):
        # Upsert по коду клиента: повторное сохранение обновляет цену, а не падает
        # на unique-ограничении.
        code = (request.data.get("client_code") or "").strip()
        existing = ClientPrice.objects.filter(client_code=code).first()
        if existing:
            serializer = self.get_serializer(existing, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data)
        return super().create(request, *args, **kwargs)


@extend_schema_view(
    list=extend_schema(
        parameters=[
            OpenApiParameter("status", OpenApiTypes.STR, enum=[s.value for s in WarehouseOrder.Status], description="Фильтр по статусу"),
            OpenApiParameter("branch", OpenApiTypes.INT, description="Филиал (менеджер/админ)"),
            OpenApiParameter("search", OpenApiTypes.STR, description="Поиск по коду клиента"),
            OpenApiParameter("active", OpenApiTypes.STR, description="1 = только активные (не Выдано/Отменена)"),
        ]
    )
)
class WarehouseOrderViewSet(viewsets.ModelViewSet):
    """Заявки на сборку (склад Loko Express).

    * Складовщик/оператор — видят только СВОЙ филиал.
    * Менеджер/админ — все заявки (фильтры status/branch/search).
    Статус меняется через action ``status`` (валидация перехода + роли).
    """

    serializer_class = WarehouseOrderSerializer
    permission_classes = [WarehouseAccess]

    def get_queryset(self):
        qs = (
            WarehouseOrder.objects
            .select_related("branch", "created_by", "assigned_to")
            .prefetch_related("items", "items__sale")
        )
        user = self.request.user
        params = self.request.query_params
        # Склад и оператор — только заявки своего филиала.
        if getattr(user, "is_warehouse", False) or getattr(user, "is_operator", False):
            qs = qs.filter(branch=user.branch) if user.branch_id else qs.none()
        else:
            branch = params.get("branch")
            if branch:
                qs = qs.filter(branch=branch)
        status = params.get("status")
        if status:
            qs = qs.filter(status=status)
        if params.get("active") in ("1", "true", "True"):
            qs = qs.exclude(status__in=[WarehouseOrder.Status.ISSUED, WarehouseOrder.Status.CANCELLED])
        # Дневная сборка: только заявки, где есть неоприходованные позиции «в поиске».
        if params.get("active_items") in ("1", "true", "True"):
            qs = qs.filter(items__status=WarehouseItem.Status.IN_SEARCH).distinct()
        search = (params.get("search") or "").strip().lower()
        if search:
            ids = [o.id for o in qs if any(search in str(c).lower() for c in (o.client_codes or []))]
            qs = qs.filter(id__in=ids)
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        # Филиал заявки: у складовщика/оператора — строго ЕГО филиал (без дефолтного
        # фолбэка). Не назначен → понятная ошибка. У менеджера/админа — выбранный.
        if getattr(user, "is_operator", False) or getattr(user, "is_warehouse", False):
            branch = user.branch
            if branch is None:
                raise serializers.ValidationError(
                    {"branch": "Вам не назначен филиал. Обратитесь к администратору — "
                               "он привяжет вас к филиалу в разделе «Пользователи»."}
                )
        else:
            branch = serializer.validated_data.get("branch") or user.branch or Branch.resolve_default()
        order = serializer.save(created_by=user, branch=branch, status=WarehouseOrder.Status.NEW)
        # Двухэтапный учёт: каждая позиция стартует «в поиске», БЕЗ денег. Продажа
        # (Sale) появится только при оприходовании конкретного кода складовщиком.
        for code in serializer.validated_data.get("client_codes", []):
            WarehouseItem.objects.create(order=order, client_code=code)

    @extend_schema(request=WarehouseStatusSerializer, responses=WarehouseOrderSerializer)
    @action(detail=True, methods=["patch", "post"], url_path="status")
    def status(self, request, pk=None):
        """Смена статуса заявки по карте TRANSITIONS.

        Складовщик ведёт весь цикл: в поиск → готова → выдано (либо отмена с
        обязательной причиной). «Взять в работу» фиксирует складовщика (assigned_to)."""
        order = self.get_object()
        ser = WarehouseStatusSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        new_status = ser.validated_data["status"]
        comment = (ser.validated_data.get("comment") or "").strip()
        user = request.user
        labels = dict(WarehouseOrder.Status.choices)

        if new_status != order.status and not order.can_transition_to(new_status):
            raise serializers.ValidationError(
                {"status": f"Недопустимый переход: {order.get_status_display()} → {labels.get(new_status, new_status)}."}
            )
        if new_status == WarehouseOrder.Status.CANCELLED and not comment:
            raise serializers.ValidationError({"comment": "Укажите причину отмены."})
        if new_status == WarehouseOrder.Status.NOT_FOUND and not comment:
            raise serializers.ValidationError({"comment": "Укажите, что не найдено / где искали."})

        order.status = new_status
        if comment:
            order.comment = comment
        if new_status == WarehouseOrder.Status.IN_PROGRESS and order.assigned_to_id is None:
            order.assigned_to = user
        order.save()
        return Response(WarehouseOrderSerializer(order).data)


@extend_schema_view(
    list=extend_schema(
        parameters=[
            OpenApiParameter(
                "status", OpenApiTypes.STR, enum=[s.value for s in WarehouseItem.Status],
                description="Фильтр по статусу позиции (напр. EVENING — вечерний допоиск)",
            ),
        ]
    )
)
class WarehouseItemViewSet(viewsets.ReadOnlyModelViewSet):
    """Позиции складской заявки (двухэтапный учёт карго).

    Складовщик оприходует ПОШТУ́ЧНО: ``receive`` (нашёл → вес → создаётся продажа)
    или ``not_found`` (нет на складе). Оператор видит свои позиции (``mine``) и может
    отправить не найденную в вечерний допоиск (``to_evening``). Область видимости
    (свой филиал / свои позиции) — в ``get_queryset``.
    """

    serializer_class = WarehouseItemSerializer
    permission_classes = [WarehouseItemAccess]

    def get_queryset(self):
        user = self.request.user
        qs = WarehouseItem.objects.select_related(
            "order", "order__branch", "order__created_by", "sale"
        )
        if getattr(user, "is_warehouse", False):
            qs = qs.filter(order__branch=user.branch) if user.branch_id else qs.none()
        elif getattr(user, "is_operator", False):
            qs = qs.filter(order__created_by=user)
        # менеджер/админ — все позиции
        status = self.request.query_params.get("status")
        if status:
            qs = qs.filter(status=status)
        return qs.order_by("id")

    @extend_schema(request=WarehouseReceiveSerializer, responses=WarehouseItemSerializer)
    @action(detail=True, methods=["post"])
    def receive(self, request, pk=None):
        """Оприходовать позицию: фактический вес → продажа по тарифу → статус FOUND."""
        item = self.get_object()
        if item.status in WarehouseItem.FINANCIAL:
            raise serializers.ValidationError({"status": "Позиция уже оприходована."})
        ser = WarehouseReceiveSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        # Заявка клиента (QR) ещё никому не закреплена — закрепляем за сотрудником
        # филиала: единственным автоматически, иначе тем, кого выбрал складовщик.
        order = item.order
        if order.created_by_id is None:
            operator = ser.validated_data.get("operator") or WarehouseOrder.resolve_operator(order.branch)
            if operator is None:
                raise serializers.ValidationError(
                    {"operator": "Выберите сотрудника, кому засчитать заявку клиента."}
                )
            if not (getattr(operator, "is_operator", False) and operator.branch_id == order.branch_id):
                raise serializers.ValidationError({"operator": "Сотрудник должен быть из филиала заявки."})
            order.assign_operator(operator)
        item.receive(
            ser.validated_data["weight_kg"], ser.validated_data["account"], by_user=request.user,
            tracking_number=ser.validated_data.get("tracking_number") or None,
        )
        return Response(WarehouseItemSerializer(item).data)

    @extend_schema(request=WarehouseNotFoundSerializer, responses=WarehouseItemSerializer)
    @action(detail=True, methods=["post"], url_path="not-found")
    def not_found(self, request, pk=None):
        """Товара нет на складе: причина → статус NOT_FOUND (продажа не создаётся)."""
        item = self.get_object()
        if item.status in WarehouseItem.FINANCIAL:
            raise serializers.ValidationError(
                {"status": "Позиция уже оприходована — нельзя пометить «не найдено»."}
            )
        ser = WarehouseNotFoundSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        item.mark_not_found(ser.validated_data["reason"], by_user=request.user)
        return Response(WarehouseItemSerializer(item).data)

    @extend_schema(request=None, responses=WarehouseItemSerializer)
    @action(detail=True, methods=["post"], url_path="to-evening")
    def to_evening(self, request, pk=None):
        """Оператор убирает не найденную позицию из чека → в вечерний допоиск склада."""
        item = self.get_object()
        if item.status != WarehouseItem.Status.NOT_FOUND:
            raise serializers.ValidationError(
                {"status": "В вечерний допоиск можно отправить только позицию «не найдено»."}
            )
        item.send_to_evening()
        return Response(WarehouseItemSerializer(item).data)

    @extend_schema(
        parameters=[OpenApiParameter("period", OpenApiTypes.STR, enum=["month", "prev", "all"],
                                     description="Период: текущий месяц (по умолчанию), прошлый месяц, всё время")],
        responses=OpenApiTypes.OBJECT,
    )
    @action(detail=False, methods=["get"])
    def mine(self, request):
        """Свои позиции сотрудника (для «Мои продажи»): текущий месяц по умолчанию,
        ``?period=prev`` — прошлый месяц, ``?period=all`` — всё время.

        Показывает статус каждого кода: в поиске / найдено (вес + сумма) / не найдено /
        вечерний допоиск. Сумма берётся с созданной при оприходовании продажи."""
        now_local = timezone.localtime(timezone.now())
        month_start = now_local.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        period = request.query_params.get("period", "month")
        qs = WarehouseItem.objects.filter(order__created_by=request.user)
        if period == "prev":
            prev_start = (month_start - timezone.timedelta(days=1)).replace(day=1)
            qs = qs.filter(created_at__gte=prev_start, created_at__lt=month_start)
        elif period != "all":
            qs = qs.filter(created_at__gte=month_start)
        qs = qs.select_related("order", "order__branch", "order__created_by", "sale").order_by("-created_at", "-id")
        data = WarehouseItemSerializer(qs, many=True).data
        return Response({"count": qs.count(), "period": period, "results": data})

    @extend_schema(parameters=[OpenApiParameter("branch", OpenApiTypes.INT)], responses=OpenApiTypes.OBJECT)
    @action(detail=False, methods=["get"])
    def operators(self, request):
        """Сотрудники филиала — пикер «кому засчитать» для заявок клиента (QR).
        Складовщик получает свой филиал; менеджер/админ — ``?branch=``."""
        user = request.user
        if getattr(user, "is_warehouse", False):
            branch = user.branch
        else:
            raw = request.query_params.get("branch")
            branch = Branch.objects.filter(pk=raw).first() if raw else None
        qs = WarehouseOrder.branch_operators(branch)
        return Response([{"id": u.id, "name": u.get_full_name() or u.username} for u in qs])

    @extend_schema(responses=OpenApiTypes.OBJECT)
    @action(detail=False, methods=["get"])
    def accounts(self, request):
        """Пикер счёта зачисления (нал/безнал) для оприходования — только Express в сомах."""
        qs = Account.objects.filter(
            module="EXPRESS", currency="KGS", is_active=True
        ).order_by("name")
        return Response([{"id": a.id, "name": a.name, "kind": a.kind} for a in qs])


# ---------------------------------------------------------------------------
# Публичные эндпоинты клиента (QR-страница, БЕЗ входа). Клиент узнаётся по
# телефону; заявка попадает складовщику филиала так же, как «Новая продажа».
# ---------------------------------------------------------------------------

@extend_schema(responses=OpenApiTypes.OBJECT, tags=["public"])
@api_view(["GET"])
@permission_classes([AllowAny])
@throttle_classes([PublicReadThrottle])
def public_branches(request):
    """Список филиалов для клиентской страницы (id/name активных). Публично."""
    qs = Branch.objects.filter(is_active=True).order_by("name")
    return Response([{"id": b.id, "name": b.name} for b in qs])


@extend_schema(request=PublicIntakeSerializer, responses=OpenApiTypes.OBJECT, tags=["public"])
@api_view(["POST"])
@permission_classes([AllowAny])
@throttle_classes([PublicWriteThrottle])
def public_intake(request):
    """Самозапись клиента по QR: телефон + имя + коды → заявка складу (позиции «в поиске»).

    Клиент регистрируется/находится по телефону; продажа НЕ создаётся (родится при
    оприходовании складом). Филиал берётся из QR."""
    ser = PublicIntakeSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    d = ser.validated_data
    client = Client.get_or_register(d["phone"], d.get("name", ""))
    # Заявка сразу закрепляется за сотрудником филиала, если он один: он увидит коды
    # в «Моих продажах» ещё до оприходования. Иначе выбор — на складе при приёме.
    order = WarehouseOrder.objects.create(
        branch=d["branch"], client=client, created_by=WarehouseOrder.resolve_operator(d["branch"]),
        status=WarehouseOrder.Status.NEW, client_codes=d["client_codes"],
    )
    for code in d["client_codes"]:
        WarehouseItem.objects.create(order=order, client_code=code)
    return Response(
        {"ok": True, "client_name": client.name, "codes": d["client_codes"]},
        status=201,
    )


@extend_schema(
    parameters=[OpenApiParameter("phone", OpenApiTypes.STR, description="Телефон клиента")],
    responses=OpenApiTypes.OBJECT, tags=["public"],
)
@api_view(["GET"])
@permission_classes([AllowAny])
@throttle_classes([PublicReadThrottle])
def public_track(request):
    """Трекинг по телефону: статусы/цены кодов клиента + бонус за вес. Публично.

    Throttled: the response reveals a client's name and parcels for any phone
    number, so the per-IP ``public_read`` cap is what stops mass phone-number
    enumeration from harvesting client data."""
    phone = Client.normalize_phone(request.query_params.get("phone"))
    if len(phone) < 6:
        return Response({"found": False, "reason": "no_phone"})
    client = Client.objects.filter(phone=phone).first()
    if client is None:
        return Response({"found": False})
    items = list(
        WarehouseItem.objects
        .filter(order__client=client)
        .select_related("order", "order__branch", "sale", "found_by")
        .order_by("-id")
    )
    total_kg = sum((i.weight_kg or ZERO) for i in items if i.status in WarehouseItem.FINANCIAL)
    free_kg = int(total_kg // 20) * Decimal("0.5")
    # Сотрудники, оприходовавшие груз клиента — их можно оценить (звёзды).
    staff = {}
    for it in items:
        u = it.found_by
        if u and it.status in WarehouseItem.FINANCIAL and u.id not in staff:
            staff[u.id] = {"id": u.id, "name": u.get_full_name() or u.username, "role": u.get_role_display(), "my_stars": 0}
    if staff:
        given = EmployeeRating.objects.filter(client=client, employee_id__in=staff.keys())
        for r in given:
            staff[r.employee_id]["my_stars"] = r.stars
    return Response({
        "found": True,
        "client": {"name": client.name, "phone": client.phone},
        "bonus": {"total_kg": str(total_kg), "free_kg": str(free_kg)},
        "items": WarehouseItemSerializer(items, many=True).data,
        "staff": list(staff.values()),
    })


@extend_schema(request=PublicRateSerializer, responses=OpenApiTypes.OBJECT, tags=["public"])
@api_view(["POST"])
@permission_classes([AllowAny])
@throttle_classes([PublicWriteThrottle])
def public_rate(request):
    """Клиент ставит звёзды сотруднику (1–5). Оценить можно только того, кто реально
    оприходовал груз этого клиента — иначе оценки можно было бы накрутить."""
    ser = PublicRateSerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    d = ser.validated_data
    client = Client.objects.filter(phone=Client.normalize_phone(d["phone"])).first()
    if client is None:
        raise serializers.ValidationError({"phone": "Клиент не найден — сначала сдайте коды."})
    served = WarehouseItem.objects.filter(
        order__client=client, found_by_id=d["employee"], status__in=WarehouseItem.FINANCIAL,
    ).exists()
    if not served:
        raise serializers.ValidationError({"employee": "Этот сотрудник не обслуживал ваш груз."})
    from django.contrib.auth import get_user_model
    emp = get_user_model().objects.get(id=d["employee"])
    EmployeeRating.objects.update_or_create(
        employee=emp, client=client, defaults={"stars": d["stars"]},
    )
    return Response({"ok": True, "employee": emp.id, "stars": d["stars"]})


@extend_schema_view(
    list=extend_schema(
        parameters=[OpenApiParameter("search", OpenApiTypes.STR, description="Поиск по имени/телефону")]
    )
)
class ClientViewSet(viewsets.ReadOnlyModelViewSet):
    """Клиенты (CRM) — имя, телефон, история заказов (кг/сумма). Кассир/админ."""

    serializer_class = ClientSerializer
    permission_classes = [DenyOperatorOrDirector]

    def get_queryset(self):
        qs = Client.objects.all()
        search = (self.request.query_params.get("search") or "").strip()
        if search:
            digits = Client.normalize_phone(search)
            cond = Q(name__icontains=search)
            if digits:
                cond |= Q(phone__icontains=digits)
            qs = qs.filter(cond)
        return qs


# ---------------------------------------------------------------------------
# Директор: «процесс работы» (прозрачность склада) и остаток веса на складе.
# ---------------------------------------------------------------------------

def _date_param(request, name):
    from datetime import date as _date
    raw = request.query_params.get(name)
    if not raw:
        return None
    try:
        return _date.fromisoformat(raw)
    except ValueError:
        raise serializers.ValidationError({name: "Ожидается дата YYYY-MM-DD."})


@extend_schema(
    parameters=[
        OpenApiParameter("from", OpenApiTypes.DATE, description="Начало периода (по умолчанию сегодня)"),
        OpenApiParameter("to", OpenApiTypes.DATE, description="Конец периода (по умолчанию сегодня)"),
        OpenApiParameter("branch", OpenApiTypes.INT, description="Филиал (пусто — все)"),
    ],
    responses=OpenApiTypes.OBJECT, tags=["reports"],
)
@api_view(["GET"])
@permission_classes([WorkflowAccess])
def workflow_report(request):
    """«Процесс работы» для директора: кто из сотрудников какие заказы обрабатывает
    (заявки, оприходованные позиции, кг, сом, не найдено), что сейчас в работе и что
    осталось на вечерний допоиск. Read-only."""
    branch = request.query_params.get("branch") or None
    return Response(build_workflow(_date_param(request, "from"), _date_param(request, "to"), int(branch) if branch else None))


@extend_schema_view(
    list=extend_schema(parameters=[OpenApiParameter("branch", OpenApiTypes.INT, description="Филиал")]),
)
class WarehouseStockViewSet(viewsets.ModelViewSet):
    """Приходы веса на склад (кг) — ведёт директор; расход считается из продаж.

    ``summary/?branch=`` — остаток, итоги и дневная лента (приход/расход/остаток)."""

    serializer_class = WarehouseStockSerializer
    permission_classes = [StockAccess]
    http_method_names = ["get", "post", "delete", "head", "options"]

    def get_queryset(self):
        qs = WarehouseStock.objects.select_related("branch", "created_by")
        branch = self.request.query_params.get("branch")
        if branch:
            qs = qs.filter(branch_id=branch)
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def perform_destroy(self, instance):
        user = self.request.user
        if not (user.is_admin or instance.created_by_id == user.id):
            raise serializers.ValidationError({"detail": "Удалить запись может её автор или администратор."})
        instance.delete()

    @extend_schema(parameters=[OpenApiParameter("branch", OpenApiTypes.INT, required=True)], responses=OpenApiTypes.OBJECT)
    @action(detail=False, methods=["get"])
    def summary(self, request):
        """Остаток кг на складе филиала: Σ приходов − Σ веса продаж с первой записи."""
        user = request.user
        if getattr(user, "is_warehouse", False):
            # Складовщик видит только свой филиал, параметр игнорируется.
            if not user.branch_id:
                raise serializers.ValidationError({"branch": "Вам не назначен филиал."})
            return Response(build_stock(user.branch_id))
        branch = request.query_params.get("branch")
        if not branch:
            b = Branch.resolve_default()
            if b is None:
                return Response(build_stock(None))
            branch = b.id
        return Response(build_stock(int(branch)))

    @extend_schema(responses=OpenApiTypes.OBJECT)
    @action(detail=False, methods=["get"])
    def branches(self, request):
        """Пикер филиалов для страницы остатка (директору обычные /branches/ закрыты)."""
        qs = Branch.objects.filter(is_active=True).order_by("name")
        return Response([{"id": b.id, "name": b.name, "is_default": b.is_default} for b in qs])
