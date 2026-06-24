"""Импорт реальных данных Loko Express из «Финансовый_учет_Локо…xlsx».

Листы:
  «10. Приход»  → продажи (Sale, режим DIRECT): Сумма начисления → выручка (ОПиУ),
                  Сумма оплаты → приток (ОДДС), разница → дебиторка.
  «11. Расход»  → расходы (Expense, OpEx): Сумма начисления/оплаты, статья.

Метод оплаты → счёт: Оптима/Мбанк/Наличные; всё остальное
(«Банк не указан», «Смешанная оплата», «Не оплачено») → счёт «Банк не указан».
"""

from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from finance.models import Account, AppSettings, Currency, Expense, ExpenseCategory, Module, OpexArticle
from express.models import Sale

DEFAULT_PATH = str(Path.home() / "Downloads" / "Финансовый_учет_Локо_приход_расход_ОПиУ_ОДДС.xlsx")
D = lambda x: Decimal(str(x or 0))


def _method_account(method, accounts):
    m = (method or "").strip().lower()
    if "оптима" in m:
        return accounts["Оптима Банк"]
    if "мбанк" in m or m == "мб":
        return accounts["МБанк"]
    if "налич" in m:
        return accounts["Наличные"]
    return accounts["Банк не указан"]


def _to_date(value):
    return value.date() if hasattr(value, "date") else value


def _hidx(ws, row, names):
    hdr = [(str(c).strip() if c else "") for c in next(ws.iter_rows(min_row=row, max_row=row, values_only=True))]
    return {n: hdr.index(n) for n in names if n in hdr}


class Command(BaseCommand):
    help = "Импорт реальных продаж и расходов Loko Express из Excel."

    def add_arguments(self, parser):
        parser.add_argument("--path", default=DEFAULT_PATH, help="Путь к xlsx файлу Express")

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("Установите openpyxl: pip install openpyxl")

        path = options["path"]
        if not Path(path).exists():
            raise CommandError(f"Файл не найден: {path}")

        cfg = AppSettings.load()
        wb = openpyxl.load_workbook(path, data_only=True)

        # Счета Express (включая «Банк не указан» для нераспознанных оплат).
        names = ["Наличные", "Оптима Банк", "МБанк", "Банк не указан"]
        accounts = {}
        for name in names:
            accounts[name], _ = Account.objects.get_or_create(
                name=name,
                defaults={
                    "kind": Account.Kind.CASH if name == "Наличные" else Account.Kind.BANK,
                    "currency": Currency.KGS,
                    "module": Module.EXPRESS,
                },
            )

        acc_ids = [a.id for a in accounts.values()]
        Sale.objects.filter(account_id__in=acc_ids).delete()
        Expense.objects.filter(account_id__in=acc_ids).delete()

        snap = dict(
            price_per_kg_usd=cfg.price_per_kg_usd,
            usd_rate_som=cfg.usd_rate_som,
            cost_per_kg_som=cfg.base_cost_per_kg_som,
        )

        # --- Приход → Sale ---
        ws = wb["10. Приход"]
        I = _hidx(ws, 4, ["Дата операции", "Дата оплаты", "Сумма начисления", "Сумма оплаты",
                          "Метод оплаты", "Клиент/контрагент", "Статья/услуга"])
        sales = []
        for r in ws.iter_rows(min_row=5, values_only=True):
            if r[0] is None:
                continue
            op_date = _to_date(r[I["Дата операции"]])
            if not op_date:
                continue
            pay_date = _to_date(r[I["Дата оплаты"]]) or op_date
            price = D(r[I["Сумма начисления"]])
            paid = D(r[I["Сумма оплаты"]])
            acc = _method_account(r[I["Метод оплаты"]], accounts)
            sales.append(Sale(
                client_code=str(r[I["Клиент/контрагент"]] or "—")[:120],
                amount_mode=Sale.AmountMode.DIRECT,
                weight_kg=None, places=1, account=acc,
                price_som=price, paid_som=paid, cost_som=Decimal("0"), margin_som=price,
                date=op_date, payment_date=pay_date, **snap,
            ))
        Sale.objects.bulk_create(sales, batch_size=500)

        # --- Расход → Expense (всё операционное) ---
        ws = wb["11. Расход"]
        I = _hidx(ws, 4, ["Дата операции", "Дата оплаты", "Сумма начисления", "Сумма оплаты",
                          "Метод оплаты", "Статья расхода", "Комментарий"])
        expenses = []
        for r in ws.iter_rows(min_row=5, values_only=True):
            if r[0] is None:
                continue
            op_date = _to_date(r[I["Дата операции"]])
            if not op_date:
                continue
            pay_date = _to_date(r[I["Дата оплаты"]]) or op_date
            article = str(r[I["Статья расхода"]] or "")
            desc = article
            if r[I.get("Комментарий", -1)] if "Комментарий" in I else None:
                desc = f"{article}: {r[I['Комментарий']]}"
            expenses.append(Expense(
                account=_method_account(r[I["Метод оплаты"]], accounts),
                category=ExpenseCategory.OPEX, opex_article=OpexArticle.OTHER,
                amount=D(r[I["Сумма начисления"]]), paid_amount=D(r[I["Сумма оплаты"]]),
                description=desc[:500], date=op_date, payment_date=pay_date,
            ))
        Expense.objects.bulk_create(expenses, batch_size=500)

        self.stdout.write(self.style.SUCCESS(
            f"✔ Импортировано продаж: {len(sales)}, расходов: {len(expenses)}"
        ))
        self.stdout.write(self.style.SUCCESS("Готово. Данные Loko Express загружены."))
